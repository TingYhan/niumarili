#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
分析 Excel 文件的结构和格式细节
"""

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from collections import defaultdict
import json

def analyze_excel(file_path):
    """分析 Excel 文件的详细信息"""
    
    wb = load_workbook(file_path)
    
    print("=" * 80)
    print("EXCEL 文件分析报告")
    print("=" * 80)
    print(f"\n文件路径: {file_path}")
    print(f"工作簿名称: {wb.title}")
    print(f"工作表数量: {len(wb.sheetnames)}")
    print(f"工作表列表: {wb.sheetnames}\n")
    
    # 分析每个工作表
    for sheet_idx, sheet_name in enumerate(wb.sheetnames):
        ws = wb[sheet_name]
        print("\n" + "=" * 80)
        print(f"工作表 {sheet_idx + 1}: {sheet_name}")
        print("=" * 80)
        
        # 基本信息
        print(f"\n【基本结构信息】")
        print(f"最大行号: {ws.max_row}")
        print(f"最大列号: {ws.max_column}")
        print(f"最大列字母: {get_column_letter(ws.max_column)}")
        
        # 合并单元格
        print(f"\n【合并单元格】")
        if ws.merged_cells:
            print(f"合并单元格数量: {len(ws.merged_cells)}")
            for merged_range in ws.merged_cells.ranges:
                print(f"  {merged_range}")
        else:
            print("无合并单元格")
        
        # 冻结窗格
        print(f"\n【冻结窗格】")
        if ws.freeze_panes:
            print(f"冻结窗格: {ws.freeze_panes}")
        else:
            print("无冻结窗格")
        
        # 页面设置
        print(f"\n【页面设置】")
        print(f"页面纸张大小: {ws.page_setup.paperSize}")
        print(f"页面方向: {ws.page_setup.orientation}")
        if ws.page_margins:
            print(f"上边距: {ws.page_margins.top}")
            print(f"下边距: {ws.page_margins.bottom}")
            print(f"左边距: {ws.page_margins.left}")
            print(f"右边距: {ws.page_margins.right}")
            print(f"页眉高: {ws.page_margins.header}")
            print(f"页脚高: {ws.page_margins.footer}")
        
        # 列宽信息
        print(f"\n【列宽信息】")
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            col_width = ws.column_dimensions[col_letter].width
            if col_width:
                print(f"  列 {col_letter}: 宽度 = {col_width}")
        
        # 行高信息
        print(f"\n【行高信息（仅显示设置了行高的行）】")
        for row_idx in range(1, min(ws.max_row + 1, 50)):  # 只显示前50行
            row_height = ws.row_dimensions[row_idx].height
            if row_height:
                print(f"  行 {row_idx}: 高度 = {row_height}")
        if ws.max_row > 50:
            print(f"  ... (总共 {ws.max_row} 行，仅显示前50行)")
        
        # 样式和格式详情（前20行）
        print(f"\n【单元格样式详情（前20行）】")
        style_details = defaultdict(list)
        
        for row_idx in range(1, min(ws.max_row + 1, 21)):
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                col_letter = get_column_letter(col_idx)
                cell_ref = f"{col_letter}{row_idx}"
                
                # 收集样式信息
                style_info = {
                    "cell": cell_ref,
                    "value": str(cell.value)[:30] if cell.value else None,
                    "font": {
                        "name": cell.font.name,
                        "size": cell.font.size,
                        "bold": cell.font.bold,
                        "italic": cell.font.italic,
                        "color": str(cell.font.color),
                    } if cell.font else None,
                    "fill": {
                        "pattern_type": cell.fill.patternType,
                        "fg_color": str(cell.fill.fgColor) if cell.fill else None,
                    } if cell.fill else None,
                    "alignment": {
                        "horizontal": cell.alignment.horizontal,
                        "vertical": cell.alignment.vertical,
                        "wrap_text": cell.alignment.wrap_text,
                    } if cell.alignment else None,
                    "border": {
                        "left": str(cell.border.left) if cell.border.left else None,
                        "right": str(cell.border.right) if cell.border.right else None,
                        "top": str(cell.border.top) if cell.border.top else None,
                        "bottom": str(cell.border.bottom) if cell.border.bottom else None,
                    } if cell.border else None,
                    "number_format": cell.number_format,
                }
                style_details[cell_ref] = style_info
        
        # 打印样式详情
        for cell_ref in sorted(style_details.keys(), key=lambda x: (int(x[1:]), x[0])):
            info = style_details[cell_ref]
            print(f"\n  {cell_ref}:")
            print(f"    值: {info['value']}")
            if info['font']:
                print(f"    字体: {info['font']['name']}, 大小: {info['font']['size']}, 粗体: {info['font']['bold']}")
            if info['alignment']:
                print(f"    对齐: 水平={info['alignment']['horizontal']}, 竖直={info['alignment']['vertical']}, 换行={info['alignment']['wrap_text']}")
            if info['border'] and any(info['border'].values()):
                print(f"    边框: 左={info['border']['left']}, 右={info['border']['right']}, 上={info['border']['top']}, 下={info['border']['bottom']}")
            if info['fill'] and info['fill']['fg_color'] and info['fill']['fg_color'] != 'None':
                print(f"    填充: {info['fill']['fg_color']}")
        
        # 表格数据区域检测
        print(f"\n【表格数据区域】")
        # 找出有内容的区域
        data_rows = []
        for row_idx in range(1, ws.max_row + 1):
            has_data = False
            for col_idx in range(1, ws.max_column + 1):
                if ws.cell(row=row_idx, column=col_idx).value is not None:
                    has_data = True
                    break
            if has_data:
                data_rows.append(row_idx)
        
        if data_rows:
            print(f"有数据的行范围: {min(data_rows)} - {max(data_rows)}")
            
            # 检查是否存在多个表格（通过空行分隔）
            tables = []
            current_table = [data_rows[0]]
            for i in range(1, len(data_rows)):
                if data_rows[i] - data_rows[i-1] == 1:
                    current_table.append(data_rows[i])
                else:
                    tables.append(current_table)
                    current_table = [data_rows[i]]
            tables.append(current_table)
            
            print(f"检测到的表格数量: {len(tables)}")
            for table_idx, table in enumerate(tables):
                print(f"  表格 {table_idx + 1}: 行 {table[0]} - {table[-1]} (共 {len(table)} 行)")
        
        # 打印完整内容（表格形式）
        print(f"\n【完整单元格内容】")
        print("(仅显示前30行)")
        print("-" * 120)
        header_row = []
        for col_idx in range(1, ws.max_column + 1):
            header_row.append(f"{get_column_letter(col_idx)}".center(15))
        print("|" + "|".join(header_row) + "|")
        print("-" * 120)
        
        for row_idx in range(1, min(ws.max_row + 1, 31)):
            row_content = []
            for col_idx in range(1, ws.max_column + 1):
                value = ws.cell(row=row_idx, column=col_idx).value
                if value is None:
                    row_content.append("".center(15))
                else:
                    content_str = str(value)[:15]
                    row_content.append(content_str.center(15))
            print("|" + "|".join(row_content) + "|")
        
        if ws.max_row > 30:
            print(f"\n... (共 {ws.max_row} 行，仅显示前30行)")

if __name__ == "__main__":
    file_path = r"c:\Users\23119\Desktop\fsdownload\加班申请表(1)(1).xlsx"
    analyze_excel(file_path)
